"""
Menu Digitalizer - foodora

Selgeren laster opp en meny (PDF, Word, Excel eller bilde).
Verktoyet ekstraherer rettene, normaliserer tekst, utleder allergener,
og viser alt i et redigerbart grid. Selgeren retter ved behov og laster
ned en Excel-fil i MDS-formatet, navngitt <Vendor>_<GRID>.xlsx.
"""

import io
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from extraction import extract_menu_from_files
from rules import to_title_case, to_sentence_case

# --- MDS-kolonner. Speiler malen fra MDS sitt CSV-output. -------------------
# Allergens og Growth_Plus er nye kolonner utenfor standard MDS.
MDS_COLUMNS = [
    "Title_en_NO", "Title_en_GB", "Title_zh_HK", "Title_en_US",
    "Description_en_NO", "Description_en_GB", "Description_zh_HK",
    "Description_en_US", "Description_type", "Category_ID",
    "Pre - packed", "Active", "Image_URL", "VAT_ID",
    "Variation_title_en_NO", "Variation_title_en_GB",
    "Variation_title_zh_HK", "Variation_title_en_US",
    "Price", "Remote_Code", "Container_Charge", "Choice_Groups_IDs",
    "Allergens",      # <-- ny kolonne
    "Growth_Plus",    # <-- ny kolonne: TRUE paa de to Growth+-rettene
]

# --- foodora / Delivery Hero fargepalett ------------------------------------
FOODORA_PINK = "#FF1F62"
FOODORA_PINK_DARK = "#D81B54"

# Lys/morkt tema-paletter. Selgeren bytter via knapp i sidepanelet.
THEMES = {
    "light": {
        "app_bg":      "#FFFFFF",
        "ink":         "#1A1A2E",
        "muted":       "#666874",
        "soft_bg":     "#FFF5F8",
        "border":      "#FFD6E2",
        "input_bg":    "#FFFFFF",
        "input_text":  "#1A1A2E",
        "header_bg":   "#FFFFFF",
        "header_text": "#FFFFFF",  # st.header er hvit boks i lys mode
        "logo_bg":     "#FFFFFF",
        "card_text":   "#1A1A2E",
    },
    "dark": {
        "app_bg":      "#15171F",
        "ink":         "#F5F5F7",
        "muted":       "#A8AAB8",
        "soft_bg":     "#23263A",
        "border":      "#3D4055",
        "input_bg":    "#2A2E40",
        "input_text":  "#F5F5F7",
        "header_bg":   "#15171F",
        "header_text": "#FFFFFF",
        "logo_bg":     "#FFFFFF",
        "card_text":   "#F5F5F7",
    },
}

# Marked -> sprakkode. Kun NO aktivt naa; lett aa utvide senere.
MARKETS = {
    "Norge": "NO",
    # "Storbritannia": "GB",
    # "USA": "US",
}

st.set_page_config(page_title="Menu Digitalizer - foodora",
                   page_icon="\U0001F37D", layout="wide")

# Sett standardtema (lyst) ved forste lasting.
if "theme" not in st.session_state:
    st.session_state.theme = "light"

T = THEMES[st.session_state.theme]
DARK = st.session_state.theme == "dark"

# --- Stil: foodora-tema, byttbart mellom lyst og morkt ----------------------
st.markdown(f"""
<style>
  /* Grunnlag */
  .stApp {{ background-color: {T['app_bg']}; color: {T['ink']}; }}

  /* Tekst: alle vanlige tekstelementer respekterer tema */
  .stApp, .stApp p, .stApp label, .stApp span, .stApp li,
  .stApp div[data-testid="stMarkdownContainer"],
  .stApp div[data-testid="stMarkdownContainer"] * {{
      color: {T['ink']};
  }}
  /* Overskrifter MAA staa ut, ogsaa st.subheader */
  .stApp h1, .stApp h2, .stApp h3,
  section[data-testid="stSidebar"] h1,
  section[data-testid="stSidebar"] h2,
  section[data-testid="stSidebar"] h3 {{
      color: {T['ink']} !important;
      font-weight: 700;
  }}
  /* Caption / hjelpetekst - tonet ned, men lesbar */
  .stApp small, .stApp [data-testid="stCaptionContainer"] {{
      color: {T['muted']} !important;
  }}

  /* Sidebar */
  section[data-testid="stSidebar"] {{
      background-color: {T['soft_bg']};
      border-right: 1px solid {T['border']};
  }}
  section[data-testid="stSidebar"] * {{
      color: {T['ink']};
  }}

  /* Primaerknapper */
  .stButton > button[kind="primary"],
  .stDownloadButton > button,
  .stFormSubmitButton > button {{
      background-color: {FOODORA_PINK};
      color: #FFFFFF !important;
      border: none;
      border-radius: 8px;
      font-weight: 600;
  }}
  .stButton > button[kind="primary"]:hover,
  .stDownloadButton > button:hover,
  .stFormSubmitButton > button:hover {{
      background-color: {FOODORA_PINK_DARK};
      color: #FFFFFF !important;
  }}
  /* Sekundaerknapper (inkl. tema-toggle) */
  .stButton > button:not([kind="primary"]) {{
      background-color: {T['input_bg']};
      color: {T['ink']};
      border: 1px solid {T['border']};
  }}

  /* Tekst- og tallinput, selectbox, radio, multiselect */
  .stTextInput input, .stNumberInput input, .stTextArea textarea {{
      background-color: {T['input_bg']} !important;
      color: {T['input_text']} !important;
      border: 1px solid {T['border']} !important;
  }}
  div[data-baseweb="select"] > div,
  div[data-baseweb="input"] > div,
  div[data-baseweb="textarea"] > div {{
      background-color: {T['input_bg']} !important;
      color: {T['input_text']} !important;
      border-color: {T['border']} !important;
  }}
  div[data-baseweb="select"] * {{ color: {T['input_text']} !important; }}
  /* Selectbox/Multiselect dropdown-meny */
  div[data-baseweb="popover"] li,
  div[data-baseweb="popover"] div {{
      background-color: {T['input_bg']} !important;
      color: {T['input_text']} !important;
  }}
  div[data-baseweb="popover"] li:hover {{
      background-color: {T['soft_bg']} !important;
  }}
  /* Radio-tekst */
  .stRadio label, .stRadio span {{ color: {T['ink']} !important; }}

  /* Filopplaster */
  section[data-testid="stFileUploaderDropzone"] {{
      background-color: {T['soft_bg']} !important;
      border: 1px dashed {T['border']} !important;
  }}
  section[data-testid="stFileUploaderDropzone"] * {{
      color: {T['ink']} !important;
  }}

  /* Metrikk-kort */
  div[data-testid="stMetric"] {{
      background-color: {T['soft_bg']};
      border: 1px solid {T['border']};
      border-radius: 10px;
      padding: 12px 16px;
  }}
  div[data-testid="stMetric"] * {{ color: {T['card_text']} !important; }}

  /* Tabeller (st.dataframe og data_editor) */
  .stDataFrame, [data-testid="stDataFrame"] {{
      background-color: {T['input_bg']} !important;
  }}

  /* Info/warning/error-bokser i dark mode trenger tilpasning */
  {"""
  div[data-testid="stAlertContainer"] {
      background-color: #2A2E40 !important;
  }
  div[data-testid="stAlertContainer"] * {
      color: #F5F5F7 !important;
  }
  """ if DARK else ""}

  /* Skille-linjer */
  hr {{ border-color: {T['border']} !important; }}

  /* Topp-banner */
  .md-banner {{
      background: linear-gradient(135deg, {FOODORA_PINK} 0%, {FOODORA_PINK_DARK} 100%);
      color: #FFFFFF;
      padding: 22px 28px;
      border-radius: 14px;
      margin-bottom: 18px;
      display: flex;
      align-items: center;
      gap: 22px;
  }}
  .md-banner h1 {{ color: #FFFFFF !important; margin: 0; font-size: 30px; }}
  .md-banner p  {{ color: #FFE3EC !important; margin: 6px 0 0 0; font-size: 15px; }}
  .md-logo {{
      background: {T['logo_bg']};
      border-radius: 12px;
      padding: 14px 18px;
      flex-shrink: 0;
      display: flex;
      align-items: center;
  }}
  .md-logo img {{ height: 56px; width: auto; display: block; }}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Hjelpefunksjoner
# ---------------------------------------------------------------------------

def items_to_dataframe(items):
    """Gjor raa ekstraksjon om til redigerbart grid med normalisering."""
    rows = []
    for it in items:
        title = to_title_case(str(it.get("title", "")).strip())
        desc = to_sentence_case(str(it.get("description", "")).strip())
        price = it.get("price")

        # Allergener kommer ferdig vurdert fra modellen, som ren liste.
        allergens = str(it.get("allergens", "")).strip()

        rows.append({
            "Tittel": title,
            "Beskrivelse": desc,
            "Variant": str(it.get("variation", "")).strip(),
            "Pris (NOK)": price if price is not None else 0.0,
            "Kategori": str(it.get("category", "")).strip(),
            "Allergener": allergens,
        })
    return pd.DataFrame(rows, columns=[
        "Tittel", "Beskrivelse", "Variant", "Pris (NOK)",
        "Kategori", "Allergener",
    ])


def safe_filename_part(text):
    """Gjor en tekst trygg som del av et filnavn."""
    cleaned = re.sub(r"[^\w\-]+", "_", text.strip())
    return cleaned.strip("_")


def build_export_filename(vendor, grid):
    """Bygg filnavn etter MDS-konvensjonen <Vendor>_<GRID>.xlsx."""
    v = safe_filename_part(vendor) or "Vendor"
    g = safe_filename_part(grid) or "GRID"
    return f"{v}_{g}.xlsx"


def build_mds_excel(df, market_lang="NO", growth_plus_indices=None):
    """
    Bygg en Excel-fil i MDS-formatet fra det redigerte gridet.

    growth_plus_indices: sett med rad-indekser (fra df) som skal merkes
    med TRUE i Growth_Plus-kolonnen. None = ingen.
    """
    growth_plus_indices = set(growth_plus_indices or [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft Menu - MDS"

    ws.append(MDS_COLUMNS)
    header_fill = PatternFill("solid", start_color=FOODORA_PINK.lstrip("#"))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    title_key = f"Title_en_{market_lang}"
    desc_key = f"Description_en_{market_lang}"
    var_key = f"Variation_title_en_{market_lang}"

    price_col_idx = MDS_COLUMNS.index("Price") + 1
    red_fill = PatternFill("solid", start_color="FFD6D6")

    for idx, r in df.iterrows():
        record = {c: "" for c in MDS_COLUMNS}
        record[title_key] = r["Tittel"]
        record[desc_key] = r["Beskrivelse"]
        record[var_key] = r["Variant"]
        record["Description_type"] = "VENDOR"
        record["Pre - packed"] = "FALSE"
        record["Active"] = "TRUE"
        record["Price"] = r["Pris (NOK)"]
        record["Allergens"] = r["Allergener"]
        if idx in growth_plus_indices:
            record["Growth_Plus"] = "TRUE"
        ws.append([record[c] for c in MDS_COLUMNS])

        # Mangler pris -> marker Price-cella rodt.
        if not r["Pris (NOK)"] or r["Pris (NOK)"] == 0:
            ws.cell(row=ws.max_row, column=price_col_idx).fill = red_fill

    for col_idx, name in enumerate(MDS_COLUMNS, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        if name.startswith("Description"):
            ws.column_dimensions[letter].width = 45
        elif name.startswith("Title") or name == "Allergens":
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 16

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

def _logo_data_uri():
    """Les Delivery Hero-logoen som base64 slik at den kan ligge i banneret."""
    import base64
    import os
    path = os.path.join(os.path.dirname(__file__), "dh_logo.png")
    try:
        with open(path, "rb") as f:
            b64 = base64.standard_b64encode(f.read()).decode("utf-8")
        return f"data:image/png;base64,{b64}"
    except FileNotFoundError:
        return None


_logo = _logo_data_uri()
_logo_html = (
    f'<div class="md-logo"><img src="{_logo}" alt="Delivery Hero"/></div>'
    if _logo else ""
)

st.markdown(f"""
<div class="md-banner">
  {_logo_html}
  <div>
    <h1>\U0001F37D Menu Digitalizer</h1>
    <p>Last opp en meny &rarr; verkt&oslash;yet strukturerer den &rarr;
       rediger og last ned i MDS-format.</p>
  </div>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    # --- Tema-bytter ---------------------------------------------------------
    is_dark = st.session_state.theme == "dark"
    label = "\u2600\ufe0f Lyst tema" if is_dark else "\U0001F319 M\u00f8rkt tema"
    if st.button(label, use_container_width=True, key="theme_toggle"):
        st.session_state.theme = "light" if is_dark else "dark"
        st.rerun()
    st.caption(
        "Vil du at temaet skal f\u00f8lge nettleseren automatisk? "
        "Trykk \u2630-menyen \u00f8verst til h\u00f8yre \u2192 Settings \u2192 "
        "sett Theme til Auto."
    )

    st.divider()
    st.header("Innstillinger")

    market_name = st.selectbox(
        "Marked",
        options=list(MARKETS.keys()),
        help="Bestemmer hvilken Title-/Description-kolonne i MDS-malen "
             "som fylles ut. Flere markeder kan legges til senere.",
    )
    market = MARKETS[market_name]

    st.divider()
    st.subheader("Vendor")
    vendor_name = st.text_input(
        "Vendornavn",
        placeholder="f.eks. Randis Gatekjokken",
        help="Brukes i filnavnet paa den nedlastede Excel-fila.",
    )
    grid_id = st.text_input(
        "GRID-id",
        placeholder="f.eks. a4b2",
        help="Vendor-id paa DH. Brukes i filnavnet etter "
             "MDS-konvensjonen <Vendor>_<GRID>.xlsx.",
    )

    st.divider()
    st.subheader("Prisjustering")
    adjust_mode = st.radio(
        "Type",
        options=["Ingen", "Prosent (%)", "Kroner (NOK)"],
        horizontal=False,
        help="Velg om alle priser skal \u00f8kes med en prosent eller "
             "et fast kronebel\u00f8p.",
    )
    price_pct = 0.0
    price_kr = 0.0
    if adjust_mode == "Prosent (%)":
        price_pct = st.number_input(
            "\u00d8k alle priser med (%)",
            min_value=0.0, max_value=100.0, value=0.0, step=1.0,
        )
    elif adjust_mode == "Kroner (NOK)":
        price_kr = st.number_input(
            "\u00d8k alle priser med (NOK)",
            min_value=0.0, max_value=500.0, value=0.0, step=1.0,
        )

# API-nokkelen leses fra Streamlit Secrets - aldri vist i UI.
try:
    api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
except Exception:
    api_key = ""

if "menu_df" not in st.session_state:
    st.session_state.menu_df = None

uploaded = st.file_uploader(
    "Last opp meny",
    type=["pdf", "docx", "xlsx", "xls", "jpg", "jpeg", "png"],
    accept_multiple_files=True,
    help="PDF, Word, Excel eller bilde av menyen. Du kan laste opp "
         "flere bilder av samme meny \u2013 de analyseres samlet.",
)
analyze_file = st.button("Analyser meny", type="primary",
                         disabled=not uploaded)

st.caption(
    "\U0001F4A1 Meny p\u00e5 nett? \u00c5pne siden i nettleseren, scroll "
    "helt til bunnen sl\u00e5 hele menyen er lastet, og lagre siden som "
    "PDF (Cmd+P \u2192 Lagre som PDF). Last s\u00e5 opp PDF-en her."
)

# --- Filopplasting -----------------------------------------------------------
if analyze_file and uploaded:
    with st.spinner(f"Analyserer {len(uploaded)} fil(er) \u2026"):
        try:
            files = [(f.getvalue(), f.name) for f in uploaded]
            items = extract_menu_from_files(files, api_key or None)
            if not items:
                st.warning("Fant ingen menyelementer i filen(e).")
            else:
                st.session_state.menu_df = items_to_dataframe(items)
                st.success(f"Hentet ut {len(items)} element(er) fra "
                           f"{len(uploaded)} fil(er). Rediger ved behov "
                           "under.")
        except ValueError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"Noe gikk galt under analysen: {e}")

# ---------------------------------------------------------------------------
# Redigerbart grid
# ---------------------------------------------------------------------------

if st.session_state.menu_df is not None:
    st.subheader("Rediger menyen")
    st.caption("Klikk i cellene for \u00e5 rette. Endringer lagres f\u00f8rst "
               "n\u00e5r du trykker **Lagre endringer** \u2013 da slipper du "
               "at siden oppdaterer seg etter hver lille rettelse.")

    with st.form("editor_form", clear_on_submit=False):
        edited_form = st.data_editor(
            st.session_state.menu_df,
            use_container_width=True,
            num_rows="dynamic",
            column_config={
                "Tittel": st.column_config.TextColumn(width="medium"),
                "Beskrivelse": st.column_config.TextColumn(width="large"),
                "Variant": st.column_config.TextColumn(width="small"),
                "Pris (NOK)": st.column_config.NumberColumn(
                    format="%.0f", min_value=0),
                "Kategori": st.column_config.TextColumn(width="small"),
                "Allergener": st.column_config.TextColumn(width="medium"),
            },
            key="editor",
        )
        saved = st.form_submit_button("\U0001F4BE Lagre endringer",
                                      type="primary")
        if saved:
            st.session_state.menu_df = edited_form

    # Det vi viser videre er den sist lagrede versjonen.
    edited = st.session_state.menu_df

    missing_mask = edited["Pris (NOK)"] == 0
    missing_price = int(missing_mask.sum())
    missing_allergens = int((edited["Allergener"].fillna("").str.strip()
                             == "").sum())
    c1, c2, c3 = st.columns(3)
    c1.metric("Retter", len(edited))
    c2.metric("Mangler pris", missing_price)
    c3.metric("Mangler allergener", missing_allergens)

    if missing_price:
        manglende = edited.loc[missing_mask, "Tittel"].tolist()
        liste = ", ".join(str(t) for t in manglende if str(t).strip())
        st.error(f"\u26a0\ufe0f {missing_price} rett(er) mangler pris og er "
                 f"merket r\u00f8dt i eksportfila: {liste}. "
                 "Fyll inn riktig pris f\u00f8r du laster opp menyen.")
    if missing_allergens:
        st.warning(f"{missing_allergens} rett(er) mangler allergener \u2013 "
                   "fyll inn f\u00f8r eksport.")

    # --- Forhaandsvisning av prisjustering -----------------------------------
    def _apply_adjustment(prices):
        """Returner justert prisserie basert paa valgt modus."""
        if price_pct > 0:
            return (prices * (1.0 + price_pct / 100.0)).round(0)
        if price_kr > 0:
            # Bare priser stoerre enn 0 justeres - vi vil ikke gi en
            # manglende-pris-rad et tilfeldig tall.
            return prices.where(prices == 0, prices + price_kr).round(0)
        return prices

    adjustment_active = price_pct > 0 or price_kr > 0
    if adjustment_active:
        adj_label = (f"+{price_pct:.0f}%" if price_pct > 0
                     else f"+{price_kr:.0f} kr")
        preview = edited[["Tittel", "Variant", "Pris (NOK)"]].copy()
        preview = preview.rename(columns={"Pris (NOK)": "Pris f\u00f8r"})
        preview[f"Pris etter {adj_label}"] = _apply_adjustment(
            edited["Pris (NOK)"])
        st.markdown(f"**Prisjustering: {adj_label}** \u2013 "
                    "slik blir prisene i eksportfila:")
        st.dataframe(preview, use_container_width=True, hide_index=True)

    st.divider()

    # --- Growth+ -------------------------------------------------------------
    st.subheader("Growth+")
    sold_growth = st.radio(
        "Er Growth+ solgt til denne vendoren?",
        options=["Nei", "Ja"],
        horizontal=True,
    )

    growth_indices = []
    if sold_growth == "Ja":
        # Bygg valg: "indeks: Tittel (Variant)" saa identiske titler skilles.
        def _label(idx, row):
            t = str(row["Tittel"]).strip() or "(uten tittel)"
            v = str(row["Variant"]).strip()
            return f"{t} - {v}" if v else t

        option_map = {
            _label(idx, row): idx for idx, row in edited.iterrows()
        }
        picked = st.multiselect(
            "Velg de 2 rettene som inng\u00e5r i Growth+-kampanjen",
            options=list(option_map.keys()),
            max_selections=2,
        )
        growth_indices = [option_map[p] for p in picked]
        if len(picked) != 2:
            st.info(f"Velg n\u00f8yaktig 2 retter \u2013 valgt n\u00e5: "
                    f"{len(picked)}.")
        else:
            st.success("2 retter valgt \u2013 merkes med TRUE i "
                       "Growth_Plus-kolonnen.")

    st.divider()

    # Anvend prosent-paaslag paa prisene for eksport.
    export_df = edited.copy()
    if adjustment_active:
        export_df["Pris (NOK)"] = _apply_adjustment(export_df["Pris (NOK)"])

    export_name = build_export_filename(vendor_name, grid_id)
    if not vendor_name or not grid_id:
        st.info(f"Fyll inn vendornavn og GRID-id i sidepanelet for "
                f"riktig filnavn. N\u00e5v\u00e6rende: **{export_name}**")
    else:
        st.caption(f"Fila lastes ned som: **{export_name}**")

    excel_buf = build_mds_excel(
        export_df, market_lang=market, growth_plus_indices=growth_indices)
    st.download_button(
        "\u2b07\ufe0f Last ned MDS-Excel",
        data=excel_buf,
        file_name=export_name,
        mime="application/vnd.openxmlformats-officedocument."
             "spreadsheetml.sheet",
        type="primary",
    )
